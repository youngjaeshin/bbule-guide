"""
Extract 허수아비침략자 sheet data from xlsx and save as JSON.
Source: 뿔레 전쟁 클리커 공략 모음 최종 수정일 26-2-22.xlsx
Output: web/data_scarecrow_invader.json
"""

import json
import os
import openpyxl

XLSX_PATH = "/Users/shin542/Desktop/Code/bbule/뿔레 전쟁 클리커 공략 모음 최종 수정일 26-2-22.xlsx"
OUTPUT_PATH = "/Users/shin542/Desktop/Code/bbule/web/data_scarecrow_invader.json"
SHEET_NAME = "허수아비침략자"

DRAGON_INVADER_NOTE = "드래곤 침략자: 클릭 불가 · 모든 피해를 0.1%만 받음 · 재화 보상 정보 미확인"

DRAGON_INVADER_BUFFS = {
    5: "공격 무효화 관통 +2.5%",
    10: "적들의 회피 확률 감소 +2.5%",
    15: "모든 용병의 공포 극복 확률 +2%",
    20: "적들의 흡수 확률 감소 +2%",
    25: "모든 용병의 치명타 딜레이 감소 +3%",
    30: "모든 용병의 추가 클릭 데미지 증폭 +2%",
    35: "적들의 약화 효과 감소 +2%",
    40: "적들의 둔화 효과 감소 +2%",
    45: "적들의 방어막 효과 감소 +2%",
    50: "적들의 부활 확률 감소 +2%",
    55: "모든 용병의 피격 지속시간 감소 +5%",
    60: "모든 용병의 피격 회피 확률 +1.5%",
    65: "클릭 크리티컬 확률 +2.5%",
    70: "모든 용병의 강타 배수 증폭 +2.5%",
    75: "적들의 피해 면제 효과 감소 +2%",
    80: "적들의 반사 확률 감소 +2%",
    85: "소울 클릭 배수 증폭 +2%",
    90: "모든 용병의 치명타 배수 증폭 +3.5%",
    95: "뿔레 토큰 드랍 확률 +3.5%",
    100: "뿔레 입자 드랍 확률 +3.5%",
    105: "적들의 무력화 효과 감소 +2%",
    110: "모든 용병의 봉인 무효화 확률 +2%",
    115: "클릭 데미지 +총 각성수 X 0.08%",
    120: "데미지 +총 각성수 X 0.1%",
    125: "모든 용병의 저항력 무시 +3%",
    130: "적들의 최대 체력 감소 +3%",
    135: "모든 용병의 행운 배수 증폭 +2.5%",
    140: "모든 종류의 쥬얼 드랍 확률 +30%",
    145: "모든 용병의 아이템 효과 증폭 +1%",
    150: "모든 용병의 아티팩트 효과 증폭 +1%",
}

DRAGON_INVADER_EFFECTS = {
    20: "아티팩트 제단 슬롯 +1",
    30: "아티팩트 보조 옵션 등급 보정 +1%",
    50: "아티팩트 제단 슬롯 +1",
    60: "아티팩트 메인 옵션 등급 보정 +1%",
    80: "아티팩트 제단 슬롯 +1",
    90: "아티팩트 보조 옵션 등급 보정 +1%",
    110: "아티팩트 제단 슬롯 +1",
    120: "아티팩트 메인 옵션 등급 보정 +1%",
    140: "아티팩트 제단 슬롯 +1",
    150: "아티팩트 제단 효과 +1%",
}


def cell_str(cell_value) -> str:
    """Convert cell value to stripped string; return '' for None/empty."""
    if cell_value is None:
        return ""
    return str(cell_value).strip()


def apply_manual_invader_overrides(invaders: dict) -> None:
    """Apply guide-maintained invader data that is not present in the source xlsx."""
    if not any(t["name"] == "드래곤 침략자" for t in invaders["types"]):
        invaders["types"].append({"name": "드래곤 침략자", "unlockStage": None})
    invaders["notes"] = [DRAGON_INVADER_NOTE]

    for reward in invaders["rewards"]:
        level = reward["level"]
        parts = []
        if level in DRAGON_INVADER_BUFFS:
            parts.append(DRAGON_INVADER_BUFFS[level])
        if level in DRAGON_INVADER_EFFECTS:
            parts.append(DRAGON_INVADER_EFFECTS[level])
        if parts:
            reward["드래곤"] = "\n/".join(parts)
        else:
            reward.pop("드래곤", None)


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # --- Invader types from row 3 (1-indexed) ---
    # "미래 침략자 125 고대침략자 150 차원 침략자 750 아카식 1250 -각 스테이지 클리어시 도전 가능"
    invader_types = [
        {"name": "미래의 침략자", "unlockStage": 125},
        {"name": "고대의 침략자", "unlockStage": 150},
        {"name": "차원 침략자",   "unlockStage": 750},
        {"name": "아카식 침략자", "unlockStage": 1250},
        {"name": "거울 침략자",   "unlockStage": None},
    ]

    # --- Invader rewards: rows 7-156 (1-indexed), cols 1-6 ---
    # Col1=level, Col2=미래, Col3=고대, Col4=차원, Col5=아카식, Col6=거울
    rewards = []
    for r in range(7, 157):  # rows 7 to 156 inclusive (levels 1-150)
        level_val = ws.cell(row=r, column=1).value
        if level_val is None:
            break
        try:
            level = int(level_val)
        except (ValueError, TypeError):
            break

        rewards.append({
            "level": level,
            "미래":  cell_str(ws.cell(row=r, column=2).value),
            "고대":  cell_str(ws.cell(row=r, column=3).value),
            "차원":  cell_str(ws.cell(row=r, column=4).value),
            "아카식": cell_str(ws.cell(row=r, column=5).value),
            "거울":  cell_str(ws.cell(row=r, column=6).value),
        })

    # --- Scarecrow: 뿔레 허수아비 (cols 7,8,9) and 변신수 허수아비 (cols 11,12,13) ---
    # hp_rule from row 4 (1-indexed)
    bbule_hp_rule = cell_str(ws.cell(row=4, column=7).value)
    transform_hp_rule = cell_str(ws.cell(row=4, column=11).value)

    bbule_levels = []
    transform_levels = []

    for r in range(7, 157):
        # 뿔레 허수아비
        lv_val = ws.cell(row=r, column=7).value
        if lv_val is not None and cell_str(lv_val) != "":
            bbule_levels.append({
                "level": cell_str(lv_val),
                "hp":    cell_str(ws.cell(row=r, column=8).value),
                "reward": cell_str(ws.cell(row=r, column=9).value),
            })

        # 변신수 허수아비
        lv_val2 = ws.cell(row=r, column=11).value
        if lv_val2 is not None and cell_str(lv_val2) != "":
            transform_levels.append({
                "level": cell_str(lv_val2),
                "hp":    cell_str(ws.cell(row=r, column=12).value),
                "reward": cell_str(ws.cell(row=r, column=13).value),
            })

    invaders = {
        "info": "각 스테이지 클리어 시 도전 가능",
        "types": invader_types,
        "rewards": rewards,
    }
    apply_manual_invader_overrides(invaders)

    output = {
        "invaders": invaders,
        "scarecrows": {
            "bbule": {
                "name": "뿔레 허수아비",
                "hpRule": bbule_hp_rule,
                "levels": bbule_levels,
            },
            "transform": {
                "name": "변신수 허수아비",
                "hpRule": transform_hp_rule,
                "levels": transform_levels,
            },
        },
    }

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"Saved: {OUTPUT_PATH}")
    print(f"  invader rewards: {len(rewards)} levels")
    print(f"  뿔레 허수아비 levels: {len(bbule_levels)}")
    print(f"  변신수 허수아비 levels: {len(transform_levels)}")


if __name__ == "__main__":
    main()
