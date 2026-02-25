#!/usr/bin/env python3
"""
build_mercenary_passive.py
엑셀 'new용병도감' 시트에서 영구 패시브 데이터를 추출하여
web/data_mercenaries.json에 passive 필드를 추가한다.

패시브는 스킬1(col 3) 텍스트에서 '-영구' 또는 '영구 -' 태그로 표시된 효과를 파싱.
"""

import json
import re
import openpyxl

XLSX_PATH = '뿔레 전쟁 클리커 공략 모음 최종 수정일 26-2-22.xlsx'
MERC_JSON = 'web/data_mercenaries.json'
SHEET_NAME = 'new용병도감'


def parse_passive(skill1_text: str) -> str | None:
    """스킬1 텍스트에서 영구 패시브 효과를 추출한다.

    다양한 포맷 처리:
      1) '효과-영구'  →  효과 부분 추출
      2) '영구 - 효과' →  효과 부분 추출
      3) '영구- 효과'  →  효과 부분 추출
    """
    if not skill1_text or '영구' not in skill1_text:
        return None

    text = skill1_text.strip()

    # 줄바꿈으로 분리
    lines = [l.strip() for l in text.split('\n') if l.strip()]

    passive_parts = []

    for line in lines:
        # '/' 로 스킬이름과 효과 분리
        if '/' in line:
            parts = line.split('/', 1)
            effect_part = parts[1].strip()
        else:
            effect_part = line

        if '영구' not in effect_part:
            continue

        # 패턴 1: "효과-영구" (영구가 뒤에)
        # 패턴 2: "영구 - 효과" 또는 "영구- 효과" (영구가 앞에)

        # 콤마로 분리된 개별 효과들 처리
        # "효과1, 효과2-영구" → 영구 태그 근처 효과 추출

        # 먼저 "영구 - 효과" 패턴 (영구가 앞에 오는 경우)
        m = re.search(r'영구\s*[-–]\s*(.+)', effect_part)
        if m:
            passive_parts.append(m.group(1).strip().rstrip(',').strip())
            continue

        # "효과-영구" 패턴 (영구가 뒤에 오는 경우)
        m = re.search(r'(.+?)\s*[-–]\s*영구', effect_part)
        if m:
            raw = m.group(1).strip()
            # 콤마로 구분된 경우, 마지막 효과만 패시브인 경우가 있음
            # 하지만 대부분 전체가 패시브
            passive_parts.append(raw.rstrip(',').strip())
            continue

        # "영구" 단독 (다음 줄에 효과)
        if effect_part.strip() == '영구':
            continue

    if not passive_parts:
        return None

    return ', '.join(passive_parts)


def normalize_grade(grade_str: str) -> str:
    """엑셀 grade 문자열을 정규화."""
    g = grade_str.strip().upper()
    # 공백 제거
    g = g.replace(' ', '')
    return g


def normalize_name(name_str: str) -> str:
    """엑셀 이름 문자열을 정규화."""
    n = name_str.strip()
    # O등급 이름에서 숫자 제거: "글렌-9500" → "글렌"
    n = re.sub(r'[-–]\d+$', '', n)
    # 끝의 하이픈 제거: "떡흐-" → "떡흐"
    n = n.rstrip('-–')
    # 공백 정리
    n = n.strip()
    return n


# 엑셀↔JSON 이름 매핑 (수동 보정)
NAME_OVERRIDES = {
    ('RBD', 'X'): 'RBD-719',
    ('데우스마키나', 'P'): '데우스 마키나',
    ('진 쿠카이', 'H'): '진.쿠카이',
}


def main():
    # 1. 엑셀에서 패시브 데이터 추출
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    xlsx_passives = {}  # (name, grade) → passive_text
    for row in range(2, ws.max_row + 1):
        grade_raw = str(ws.cell(row=row, column=1).value or '').strip()
        name_raw = str(ws.cell(row=row, column=2).value or '').strip()
        skill1 = str(ws.cell(row=row, column=3).value or '')

        if not name_raw or '영구' not in skill1:
            continue

        grade = normalize_grade(grade_raw)
        name = normalize_name(name_raw)
        passive = parse_passive(skill1)

        if passive:
            # 이름 오버라이드 적용
            override_key = (name, grade)
            if override_key in NAME_OVERRIDES:
                name = NAME_OVERRIDES[override_key]
            xlsx_passives[(name, grade)] = passive

    print(f"엑셀에서 추출한 패시브: {len(xlsx_passives)}건")

    # 2. data_mercenaries.json 로드
    with open(MERC_JSON, 'r', encoding='utf-8') as f:
        mercs = json.load(f)

    # 3. 매칭 및 패시브 추가
    matched = 0
    unmatched_xlsx = set(xlsx_passives.keys())

    for m in mercs:
        key = (m['name'], m['grade'])
        if key in xlsx_passives:
            m['passive'] = xlsx_passives[key]
            matched += 1
            unmatched_xlsx.discard(key)
        else:
            m['passive'] = None

    print(f"매칭 성공: {matched}/{len(xlsx_passives)}")

    if unmatched_xlsx:
        print(f"\n매칭 실패 ({len(unmatched_xlsx)}건):")
        for name, grade in sorted(unmatched_xlsx):
            print(f"  [{grade}] {name}: {xlsx_passives[(name, grade)]}")

    # 4. 저장
    with open(MERC_JSON, 'w', encoding='utf-8') as f:
        json.dump(mercs, f, ensure_ascii=False, indent=1)

    print(f"\n{MERC_JSON} 업데이트 완료 (passive 필드 추가)")

    # 5. 통계
    with_passive = sum(1 for m in mercs if m.get('passive'))
    print(f"패시브 보유 용병: {with_passive}/{len(mercs)}")


if __name__ == '__main__':
    main()
